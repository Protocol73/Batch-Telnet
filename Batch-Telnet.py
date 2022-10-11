#Python for mass edit's of telnet devices.
#With great power comes great responsibility.
#Written by Protocol73
''' 
THE TO DO LIST:
	Clean up imports
	Setup config files?
		Already using excel sheet

	setup if __main__ ?
'''
#END THE TO DO LIST
#=========================================================================
Notice = "-=-=- THIS IS BETA -=-=- \nThis is a tested version but NOT Recommended everyday use.\n"
ExcelNotice = "Do NOT have the Excel Sheet this Script will be using open while running this!\n"
FinalNotice = "\nYou should only be using this if you understand the risks of this & telnet.\n === YOU HAVE BEEN WARNED === \n"
P73c_Ver = 0.3 #for checking core imports -NYI-
StartupDelay = 5 #move to config?
#=========================================================================
import sys #for using args
import time #for time.sleep()
import os.path #for checking files/dir 
from datetime import date #inthename
from getpass import getpass #InTheName
import telnetlib #https://docs.python.org/3/library/telnetlib.html
from openpyxl import load_workbook #for xlsx files
from configparser import ConfigParser #for reading .cfg files
#=========================================================================
#Excel Stuff
exceldata = ("template.xlsx")
wb = load_workbook(exceldata)
print ("Sheets on workbook:")
print (wb.sheetnames)
ws = wb['Devices']
#=========================================================================
#Defined varibles
TelnetPrompt = 'apc>'
TelnetBatchRun = b"web -h enable\r"
reboot_cmd = b"reboot\r"
reboot_after = True
#username/password for batch
user = input("Enter Username: ")
password = getpass()
#=========================================================================
def startup():
	os.system('cls')
	print(Notice)
	print(ExcelNotice)
	print("Using File: " + exceldata)
	time.sleep(2)
	print(FinalNotice)
	print("Starting in: " + str(StartupDelay) + " Seconds")
	time.sleep(StartupDelay)
#=========================================================================

def telnetchanges(Device,reboot,countcurrent):
	HOST = Device
	try:
		tn = telnetlib.Telnet(HOST)
		tn.read_until(b"User Name :",5)
		tn.write(user.encode('ascii') + b"\r")
		if password:
		    tn.write(password.encode('ascii') + b"\r")
		logintimestart = time.time()
		tn.read_until(b"apc>",4)
		logintimestop = time.time()
		logintime = logintimestop - logintimestart
		if logintime < 4:
			print("Login success for: " + Device)
		else:
			print("login failed for: " + Device + " Timed out [4sec]")
			ws['D' + str(countcurrent)] = "Login Failed"
			ws['G' + str(countcurrent)] = "Login Failed"
			pass
		WEB_ENABLE_Start = time.time()
		tn.write(TelnetBatchRun)
		tn.read_until(b"apc>",3)
		WEB_ENABLE_Stop = time.time()
		WEB_ENABLE_Time = WEB_ENABLE_Stop - WEB_ENABLE_Start
		if WEB_ENABLE_Time > 3:
			print("Web enable on: " + Device + " Timed out. [3sec]")
			RebootNote = "G" + str(countcurrent)
			RebootNoteCurrent = ws[RebootNote].value
			ws['G' + str(countcurrent)] = "Run Failed"
			reboot = False
		else:
			print("HTTP enabled on:" + Device)
			RunNote = "D" + str(countcurrent)
			RunNoteCurrent = ws[RunNote].value
			ws['D' + str(countcurrent)] = "Run Sucess"

		if reboot is True:
			def telnetReboot(Device):
				time.sleep(0.2)
				tn.write(reboot_cmd)
				print("Sending reboot request to: " + Device)
				REBOOT_CHECK_start = time.time()
				tn.read_until(b"Enter 'YES' to continue or <ENTER> to cancel :",3)
				REBOOT_CHECK_stop = time.time()
				REBOOT_CHECK_Time = REBOOT_CHECK_stop - REBOOT_CHECK_start	
				if REBOOT_CHECK_Time > 3:
					print("\nRebooting " + Device + " failed")
					reboot_done = False
					tn.write(b"exit\n",2)
					ws['G' + str(countcurrent)] = False
					print("============== TELNET DEBUG OUTPUT ==================")
					print(tn.read_all().decode('ascii'))
				else:
					tn.write(b"yes")
					print(Device + " Reboot CMD sent: Success")
					reboot_done = True
				return reboot_done

			reboot_done = telnetReboot(Device)
			if reboot_done == True:
				ws['D' + str(countcurrent)] = "Success"
				ws['G' + str(countcurrent)] = "Reboot Sent"
			else:
				ws['D' + str(countcurrent)] = False
				ws['G' + str(countcurrent)] = "Reboot Failed"
		else:
			print("No reboot requested,exiting: " + Device)
			tn.write(b"exit\n")
	except ConnectionRefusedError:
		print("Device: " + Device + " responding to ping, but refused telnet connection.")
		DeviceNotes = "D" + str(countcurrent)
		DeviceNotesCurrent = ws[DeviceNotes].value
		ws['D' + str(countcurrent)] = "Telnet Refused"
		ws['G' + str(countcurrent)] = "Failed"
	except EOFError:
		print("Telnet timeout/disconnected on:" + Device)
		ws['D' + str(countcurrent)] = "Telnet EOF/Timeout"
		ws['G' + str(countcurrent)] = "Failed"
	finally:
		pass
#=========================================================================
def CheckOnline():
	countcurrent = 1
	stoppoint = 254 #don't have more then this on one sheet.
	while countcurrent <= stoppoint:
		countcurrent = countcurrent + 1
		apcDeviceIP = "C" + str(countcurrent)
		apcDeviceIPCurrent = ws[apcDeviceIP].value
		if apcDeviceIPCurrent != None:
			response = os.system('ping -n 2 ' + apcDeviceIPCurrent)
			if response == 0:
				print(apcDeviceIPCurrent, " :IS ACTIVE")
				ws['E' + str(countcurrent)] = True
				ws['F' + str(countcurrent)] = date.today()
				telnetchanges(apcDeviceIPCurrent,reboot_after,countcurrent)
			else:
				print(apcDeviceIPCurrent, "PING FAILED. HOST IS DOWN")
				print("Not logging into:" + apcDeviceIPCurrent + "\nDevice is not responding to ping.")
				DeviceNotes = "D" + str(countcurrent)
				DeviceNotesCurrent = ws[DeviceNotes].value
				ws['D' + str(countcurrent)] = "ICMP Timeout"
				ws['E' + str(countcurrent)] = False
				ws['G' + str(countcurrent)] = "Offline?"
		else:
			break
	print("Data & Dates will be added to sheet")
#=========================================================================
def main():
	CheckOnline()
	try:
		wb.save(exceldata)
	except PermissionError:
		print("=========================================================================")
		print("Make sure the Excel file is closed, so the script can save data.")
		input("Press Enter to continue:")
		wb.save(exceldata)
	print("\n===EXIT===")

main()
#=========================================================================
#END OF FILE